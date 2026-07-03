"""
Project Z - Roster Export API
PDF, Excel, and CSV roster export endpoints.
"""
import calendar
import csv
import io
import logging
from datetime import date, datetime
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import and_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import get_current_user, require_hr_admin
from app.database.session import get_db
from app.models.roster import AssignmentType, RosterEntry, RosterSnapshot

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/roster-exports", tags=["Roster Exports"])

ASSIGN_COLORS = {
    "DAY": {"bg": "#FEF9C3", "text": "#854D0E", "short": "D"},
    "NIGHT": {"bg": "#EDE9FE", "text": "#4C1D95", "short": "N"},
    "OFF": {"bg": "#F3F4F6", "text": "#6B7280", "short": "—"},
    "ADMIN": {"bg": "#DBEAFE", "text": "#1E40AF", "short": "A"},
    "LEAVE": {"bg": "#DCFCE7", "text": "#166534", "short": "L"},
    "ABSENT": {"bg": "#FEE2E2", "text": "#991B1B", "short": "!"},
    "HOLIDAY": {"bg": "#FCE7F3", "text": "#831843", "short": "H"},
}


def _month_date_range(year: int, month: int) -> tuple[date, date]:
    first = date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    last = date(year, month, last_day)
    return first, last


async def _load_entries(
    db: AsyncSession,
    department_id: UUID,
    year: int,
    month: int,
) -> tuple[Optional[RosterSnapshot], list[RosterEntry], list[str]]:
    snap_result = await db.execute(
        select(RosterSnapshot).where(
            and_(
                RosterSnapshot.department_id == department_id,
                RosterSnapshot.year == year,
                RosterSnapshot.month == month,
            )
        )
    )
    snap = snap_result.scalar_one_or_none()
    if not snap:
        return None, [], []

    entries_result = await db.execute(
        select(RosterEntry)
        .where(RosterEntry.snapshot_id == snap.id)
        .order_by(RosterEntry.entry_date, RosterEntry.employee_name)
    )
    entries = list(entries_result.scalars().all())

    all_days = sorted(set(e.entry_date.isoformat() for e in entries))
    return snap, entries, all_days


def _build_employee_map(entries: list[RosterEntry], days: list[str]) -> list[dict]:
    emp_map: dict[str, dict] = {}
    for e in entries:
        emp_key = str(e.employee_id)
        if emp_key not in emp_map:
            emp_map[emp_key] = {
                "id": emp_key,
                "code": e.employee_code,
                "name": e.employee_name,
                "pair_name": e.pair_name,
                "schedule": {},
            }
        emp_map[emp_key]["schedule"][e.entry_date.isoformat()] = {
            "assignment": e.assignment.value,
            "shift_start": e.shift_start,
            "shift_end": e.shift_end,
            "is_overridden": e.is_overridden,
        }
    result = sorted(emp_map.values(), key=lambda x: (x["pair_name"] or "ZZZ", x["name"]))
    return result


@router.get("/csv/{dept_id}")
async def export_csv(
    dept_id: UUID,
    year: int = Query(..., ge=2020, le=2100),
    month: int = Query(..., ge=1, le=12),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_hr_admin),
):
    snap, entries, days = await _load_entries(db, dept_id, year, month)
    if not snap or not entries:
        raise HTTPException(404, "No roster found for this period")

    employees = _build_employee_map(entries, days)

    output = io.StringIO()
    writer = csv.writer(output)

    header = ["Employee Code", "Employee Name", "Pair"] + [d[-5:] for d in days]
    writer.writerow(header)

    for emp in employees:
        row = [emp["code"], emp["name"], emp["pair_name"] or ""]
        for day in days:
            cell = emp["schedule"].get(day, {})
            row.append(cell.get("assignment", ""))
        writer.writerow(row)

    output.seek(0)
    from datetime import timezone
    filename = f"roster_{snap.department_name}_{year}-{month:02d}.csv"
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/excel/{dept_id}")
async def export_excel(
    dept_id: UUID,
    year: int = Query(..., ge=2020, le=2100),
    month: int = Query(..., ge=1, le=12),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_hr_admin),
):
    snap, entries, days = await _load_entries(db, dept_id, year, month)
    if not snap or not entries:
        raise HTTPException(404, "No roster found for this period")

    employees = _build_employee_map(entries, days)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl is not installed. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Roster {year}-{month:02d}"

    bold_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    day_headers = [d[-5:] for d in days]
    header_row = ["Employee Code", "Employee Name", "Pair"] + day_headers
    ws.append(header_row)

    for col_idx, _ in enumerate(header_row, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.row_dimensions[1].height = 28

    for emp_idx, emp in enumerate(employees, 2):
        row_data = [emp["code"], emp["name"], emp["pair_name"] or ""]
        for day in days:
            cell_data = emp["schedule"].get(day, {})
            row_data.append(cell_data.get("assignment", ""))
        ws.append(row_data)

        for col_idx in range(1, len(header_row) + 1):
            cell = ws.cell(row=emp_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[emp_idx].height = 22

        day_start_col = 4
        for d_idx, day in enumerate(days):
            cell_data = emp["schedule"].get(day, {})
            assign = cell_data.get("assignment", "")
            cfg = ASSIGN_COLORS.get(assign, ASSIGN_COLORS["OFF"])
            cell = ws.cell(row=emp_idx, column=day_start_col + d_idx)
            if assign:
                cell.fill = PatternFill(start_color=cfg["bg"][1:], end_color=cfg["bg"][1:], fill_type="solid")
                cell.font = Font(color=cfg["text"][1:], bold=True)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    for d_idx in range(len(days)):
        ws.column_dimensions[chr(68 + d_idx) if d_idx < 24 else f"A{68 + d_idx - 24}"].width = 6

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"roster_{snap.department_name}_{year}-{month:02d}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/pdf/{dept_id}")
async def export_pdf(
    dept_id: UUID,
    year: int = Query(..., ge=2020, le=2100),
    month: int = Query(..., ge=1, le=12),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_hr_admin),
):
    snap, entries, days = await _load_entries(db, dept_id, year, month)
    if not snap or not entries:
        raise HTTPException(404, "No roster found for this period")

    employees = _build_employee_map(entries, days)

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A3
        from reportlab.lib.units import mm, cm
        from reportlab.pdfgen import canvas
        from reportlab.platypus import Table, TableStyle, Paragraph, Spacer, SimpleDocTemplate
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        raise HTTPException(500, "reportlab is not installed. Install with: pip install reportlab")

    output = io.BytesIO()

    month_name = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"][month - 1]

    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A3),
        title=f"Roster - {snap.department_name} - {month_name} {year}",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "RosterTitle", parent=styles["Title"],
        fontSize=16, spaceAfter=6, spaceBefore=6,
    )
    subtitle_style = ParagraphStyle(
        "RosterSubtitle", parent=styles["Normal"],
        fontSize=10, textColor=colors.grey, spaceAfter=12,
    )

    elements = []
    elements.append(Paragraph(f"Department Roster — {snap.department_name}", title_style))
    elements.append(Paragraph(f"{month_name} {year} &nbsp;|&nbsp; Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))

    header = ["Employee", "Code", "Pair"] + [d[-2:] for d in days]
    data_rows = [header]
    day_color_map = {}

    for d_idx, day_str in enumerate(days):
        d = date.fromisoformat(day_str)
        iso_wd = d.isoweekday()
        if iso_wd in (6, 7):
            day_color_map[d_idx + 3] = colors.Color(0.95, 0.95, 0.97)

    for emp in employees:
        row = [emp["name"], emp["code"], emp["pair_name"] or ""]
        for day in days:
            cell_data = emp["schedule"].get(day, {})
            assign = cell_data.get("assignment", "")
            cfg = ASSIGN_COLORS.get(assign, ASSIGN_COLORS["OFF"])
            row.append(cfg["short"])
        data_rows.append(row)

    col_widths = [120, 50, 60] + [28] * len(days)
    page_width = landscape(A3)[0] - 40 * mm
    if sum(col_widths) > page_width:
        scale = page_width / sum(col_widths)
        col_widths = [int(w * scale) for w in col_widths]

    table = Table(data_rows, colWidths=col_widths, repeatRows=1)

    style_commands = [
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.12, 0.16, 0.22)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 1), (2, -1), 7),
        ("FONTSIZE", (3, 1), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.Color(0.82, 0.82, 0.84)),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.Color(0.97, 0.97, 0.98)]),
    ]

    for col_idx, bg_color in day_color_map.items():
        style_commands.append(("BACKGROUND", (col_idx, 0), (col_idx, -1), bg_color))

    for r_idx in range(1, len(data_rows)):
        for c_idx in range(3, len(data_rows[r_idx])):
            cell_val = data_rows[r_idx][c_idx]
            for assign_key, cfg in ASSIGN_COLORS.items():
                if cell_val == cfg["short"]:
                    hex_color = cfg["bg"][1:]
                    r = int(hex_color[0:2], 16) / 255
                    g = int(hex_color[2:4], 16) / 255
                    b = int(hex_color[4:6], 16) / 255
                    style_commands.append(
                        ("BACKGROUND", (c_idx, r_idx), (c_idx, r_idx), colors.Color(r, g, b))
                    )
                    hex_text = cfg["text"][1:]
                    tr = int(hex_text[0:2], 16) / 255
                    tg = int(hex_text[2:4], 16) / 255
                    tb = int(hex_text[4:6], 16) / 255
                    style_commands.append(
                        ("TEXTCOLOR", (c_idx, r_idx), (c_idx, r_idx), colors.Color(tr, tg, tb))
                    )
                    break

    table.setStyle(TableStyle(style_commands))
    elements.append(table)

    elements.append(Spacer(1, 12))
    legend_data = [["D=Day", "N=Night", "—=Off", "A=Admin", "L=Leave", "!=Absent", "H=Holiday"]]
    legend = Table(legend_data, colWidths=[60] * 7)
    legend.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.grey),
    ]))
    elements.append(legend)

    doc.build(elements)
    output.seek(0)

    filename = f"roster_{snap.department_name}_{year}-{month:02d}.pdf"
    return Response(
        content=output.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
