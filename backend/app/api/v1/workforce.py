"""
Project Z - Workforce Planning API
Department-centric workforce management, roster calendars, shift operations.

Endpoints:
  GET  /workforce/departments/summary              — All departments with today's stats
  GET  /workforce/departments/{dept_id}/detail      — Department detail + employee table
  GET  /workforce/departments/{dept_id}/roster      — Monthly roster grid
  GET  /workforce/employees/{emp_id}/profile        — Employee workforce profile
  GET  /workforce/employees/{emp_id}/calendar       — Monthly shift calendar
  POST /workforce/shift-change                      — Change an employee's shift
  POST /workforce/shift-swap                        — Swap shifts between two employees
  GET  /workforce/coverage                          — Current shift coverage stats
  GET  /workforce/upcoming-changes                  — Upcoming shift changes & returns
  POST /workforce/roster/export                     — Export department roster (PDF/Excel)
"""

import calendar
import io
import logging
from datetime import date, datetime, timedelta, timezone
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import and_, func, select, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.core.config import get_settings
from app.core.dependencies import get_current_user, PermissionChecker
from app.database.session import get_db
from app.models.attendance import AttendanceSession
from app.models.audit import AuditLog
from app.models.department import Department
from app.models.employee import Employee, EmployeeStatus
from app.models.enrollment_session import EnrollmentSession
from app.models.leave_request import LeaveRequest, LeaveStatus
from app.models.office import Office
from app.models.shift_assignment import EmployeeShiftAssignment
from app.models.shift_override import EmployeeShiftOverride
from app.models.shift_template import ShiftTemplate
from app.models.shift_protocol import ShiftProtocol
from app.models.expected_attendance import ExpectedAttendance

logger = logging.getLogger(__name__)
settings = get_settings()

router = APIRouter(prefix="/workforce", tags=["Workforce Planning"])


# ── Schemas ───────────────────────────────────────────────────

class ShiftChangeRequest(BaseModel):
    employee_id: UUID
    shift_template_id: UUID
    start_date: date
    end_date: date
    reason: Optional[str] = None


class ShiftSwapRequest(BaseModel):
    employee_a_id: UUID
    employee_b_id: UUID
    swap_date: date
    reason: Optional[str] = None


class RosterExportRequest(BaseModel):
    department_id: UUID
    year: int
    month: int
    format: str = "xlsx"  # xlsx or csv


# ── Helpers ───────────────────────────────────────────────────

def _calendar_days(year: int, month: int) -> list[date]:
    """Return all dates in a given month."""
    num_days = calendar.monthrange(year, month)[1]
    return [date(year, month, d) for d in range(1, num_days + 1)]


async def _resolve_shift_for_employee_on_date(
    db: AsyncSession, employee_id: UUID, target_date: date
) -> Optional[ShiftTemplate]:
    """Resolve the active shift template for an employee on a date (simplified 3-level)."""
    # Level 1: Override
    override = (await db.execute(
        select(EmployeeShiftOverride).where(
            and_(
                EmployeeShiftOverride.employee_id == employee_id,
                EmployeeShiftOverride.start_date <= target_date,
                EmployeeShiftOverride.end_date >= target_date,
            )
        ).order_by(EmployeeShiftOverride.created_at.desc()).limit(1)
    )).scalar_one_or_none()
    if override:
        tmpl = (await db.execute(
            select(ShiftTemplate).where(ShiftTemplate.id == override.shift_template_id)
        )).scalar_one_or_none()
        return tmpl

    # Level 2: Assignment
    assignment = (await db.execute(
        select(EmployeeShiftAssignment).where(
            EmployeeShiftAssignment.employee_id == employee_id
        ).order_by(EmployeeShiftAssignment.created_at.desc()).limit(1)
    )).scalar_one_or_none()
    if assignment:
        tmpl_id = assignment.resolve_template_id_for_date(target_date)
        if tmpl_id:
            return (await db.execute(
                select(ShiftTemplate).where(ShiftTemplate.id == tmpl_id)
            )).scalar_one_or_none()

    # Level 3: Department rule
    emp = (await db.execute(
        select(Employee).where(Employee.id == employee_id)
    )).scalar_one_or_none()
    if not emp or not emp.department_id:
        return None

    from app.models.dept_shift_rule import DepartmentShiftRule
    dept_rule = (await db.execute(
        select(DepartmentShiftRule).where(
            and_(
                DepartmentShiftRule.department_id == emp.department_id,
                DepartmentShiftRule.effective_from <= target_date,
                (DepartmentShiftRule.effective_to.is_(None) | (DepartmentShiftRule.effective_to >= target_date)),
            )
        ).order_by(DepartmentShiftRule.effective_from.desc()).limit(1)
    )).scalar_one_or_none()
    if dept_rule:
        return (await db.execute(
            select(ShiftTemplate).where(ShiftTemplate.id == dept_rule.shift_template_id)
        )).scalar_one_or_none()

    return None


async def _get_shift_label(
    db: AsyncSession, employee_id: UUID, target_date: date
) -> str:
    """Return a display label: 'D', 'N', 'OFF', '—'."""
    tmpl = await _resolve_shift_for_employee_on_date(db, employee_id, target_date)
    if tmpl is None:
        return "—"
    if tmpl.code.upper() == "OFF":
        return "OFF"
    if tmpl.is_overnight:
        return "N"
    # Classify by start time: before noon = Day, after noon = Night
    if tmpl.start_time.hour < 12:
        return "D"
    return "N"


# ── Department Summaries ──────────────────────────────────────

@router.get("/departments/summary", dependencies=[Depends(PermissionChecker("employee:view"))])
async def department_summaries(
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """All departments with today's workforce stats."""
    today = date.today()
    result = await db.execute(
        select(Department).options(
            joinedload(Department.office),
            joinedload(Department.shift_protocol),
        ).where(Department.is_active == True).order_by(Department.name)
    )
    departments = result.unique().scalars().all()

    # Get enrolled employee IDs for filtering
    enrolled_result = await db.execute(
        select(Employee.id)
        .join(EnrollmentSession, EnrollmentSession.employee_id == Employee.id)
        .where(Employee.status == EmployeeStatus.ACTIVE)
        .distinct()
    )
    enrolled_ids = set(row[0] for row in enrolled_result.all())

    summaries = []
    for dept in departments:
        # Employee count (only enrolled)
        emp_count = (await db.execute(
            select(func.count()).select_from(Employee).where(
                and_(
                    Employee.department_id == dept.id,
                    Employee.status == EmployeeStatus.ACTIVE,
                    Employee.id.in_(enrolled_ids) if enrolled_ids else True,
                )
            )
        )).scalar_one()

        # Present today (from attendance_sessions, only enrolled)
        present_filters = [
            Employee.department_id == dept.id,
            AttendanceSession.date == today,
            AttendanceSession.status.in_(["present", "late", "early_arrival", "on_time", "unscheduled_attendance"]),
        ]
        if enrolled_ids:
            present_filters.append(Employee.id.in_(enrolled_ids))
        present_count = (await db.execute(
            select(func.count()).select_from(AttendanceSession)
            .join(Employee, Employee.id == AttendanceSession.employee_id)
            .where(and_(*present_filters))
        )).scalar_one()

        # Late today (only enrolled)
        late_filters = [
            Employee.department_id == dept.id,
            AttendanceSession.date == today,
            AttendanceSession.status == "late",
        ]
        if enrolled_ids:
            late_filters.append(Employee.id.in_(enrolled_ids))
        late_count = (await db.execute(
            select(func.count()).select_from(AttendanceSession)
            .join(Employee, Employee.id == AttendanceSession.employee_id)
            .where(and_(*late_filters))
        )).scalar_one()

        # On leave today
        leave_count = (await db.execute(
            select(func.count()).select_from(LeaveRequest)
            .where(
                and_(
                    LeaveRequest.employee_id.in_(
                        select(Employee.id).where(Employee.department_id == dept.id)
                    ),
                    LeaveRequest.status == LeaveStatus.APPROVED,
                    LeaveRequest.start_date <= today,
                    LeaveRequest.end_date >= today,
                )
            )
        )).scalar_one()

        # Day/Night shift staff count (only enrolled)
        day_count = 0
        night_count = 0
        emp_query = select(Employee.id).where(
            and_(Employee.department_id == dept.id, Employee.status == EmployeeStatus.ACTIVE)
        )
        if enrolled_ids:
            emp_query = emp_query.where(Employee.id.in_(enrolled_ids))
        employees = (await db.execute(emp_query)).scalars().all()
        for emp_id in employees:
            lbl = await _get_shift_label(db, emp_id, today)
            if lbl == "D":
                day_count += 1
            elif lbl == "N":
                night_count += 1

        absent_count = max(0, emp_count - present_count - leave_count)

        summaries.append({
            "department_id": str(dept.id),
            "department_name": dept.name,
            "department_code": dept.code,
            "office_name": dept.office.name if dept.office else None,
            "shift_protocol_name": dept.shift_protocol.name if dept.shift_protocol else None,
            "head_name": dept.head_name,
            "total_employees": emp_count,
            "present_today": present_count,
            "late_today": late_count,
            "absent_today": absent_count,
            "on_leave": leave_count,
            "day_shift_staff": day_count,
            "night_shift_staff": night_count,
        })

    return {"departments": summaries, "total_departments": len(summaries)}


# ── Department Detail ─────────────────────────────────────────

@router.get("/departments/{dept_id}/detail", dependencies=[Depends(PermissionChecker("employee:view"))])
async def department_detail(
    dept_id: UUID,
    status_filter: Optional[str] = Query(None),
    shift_filter: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Department workforce view with employee table."""
    today = date.today()

    dept = (await db.execute(
        select(Department).options(
            joinedload(Department.office),
            joinedload(Department.shift_protocol),
        ).where(Department.id == dept_id)
    )).scalar_one_or_none()
    if not dept:
        raise HTTPException(404, "Department not found")

    # Fetch employees (only enrolled)
    enrolled_subq = (
        select(EnrollmentSession.employee_id)
        .distinct()
        .where(EnrollmentSession.employee_id == Employee.id)
    ).correlate(Employee).exists()
    query = select(Employee).where(
        and_(
            Employee.department_id == dept_id,
            Employee.status == EmployeeStatus.ACTIVE,
            enrolled_subq,
        )
    )
    if search:
        search_pattern = f"%{search}%"
        query = query.where(
            Employee.full_name.ilike(search_pattern) | Employee.employee_code.ilike(search_pattern)
        )
    query = query.order_by(Employee.full_name)
    employees = (await db.execute(query)).scalars().all()

    # Build employee rows with live status
    employee_rows = []
    for emp in employees:
        # Today's session
        session = (await db.execute(
            select(AttendanceSession).where(
                and_(
                    AttendanceSession.employee_id == emp.id,
                    AttendanceSession.date == today,
                )
            )
        )).scalar_one_or_none()

        # Resolve current shift
        shift_label = await _get_shift_label(db, emp.id, today)
        shifttmpl = await _resolve_shift_for_employee_on_date(db, emp.id, today)
        shift_name = shifttmpl.name if shifttmpl else "Unassigned"

        # Resolve individual protocol: check if employee has their own override or assignment
        has_individual_override = False
        emp_protocol_name = dept.shift_protocol.name if dept.shift_protocol else None

        # Level 1: Override
        override = (await db.execute(
            select(EmployeeShiftOverride).where(
                and_(
                    EmployeeShiftOverride.employee_id == emp.id,
                    EmployeeShiftOverride.start_date <= today,
                    EmployeeShiftOverride.end_date >= today,
                )
            ).limit(1)
        )).scalar_one_or_none()
        if override:
            has_individual_override = True
            override_tmpl = (await db.execute(
                select(ShiftTemplate).where(ShiftTemplate.id == override.shift_template_id)
            )).scalar_one_or_none()
            if override_tmpl:
                emp_protocol_name = f"{override_tmpl.name} (Override)"

        # Level 2: Individual assignment
        if not has_individual_override:
            assignment = (await db.execute(
                select(EmployeeShiftAssignment).where(
                    EmployeeShiftAssignment.employee_id == emp.id
                ).limit(1)
            )).scalar_one_or_none()
            if assignment:
                has_individual_override = True
                assign_tmpl_id = assignment.resolve_template_id_for_date(today)
                if assign_tmpl_id:
                    assign_tmpl = (await db.execute(
                        select(ShiftTemplate).where(ShiftTemplate.id == assign_tmpl_id)
                    )).scalar_one_or_none()
                    if assign_tmpl:
                        emp_protocol_name = f"{assign_tmpl.name} (Assigned)"
                else:
                    emp_protocol_name = "Custom Rotation"

        # Next shift (tomorrow)
        tomorrow = today + timedelta(days=1)
        next_shift_label = await _get_shift_label(db, emp.id, tomorrow)
        next_shifttmpl = await _resolve_shift_for_employee_on_date(db, emp.id, tomorrow)
        next_shift_name = next_shifttmpl.name if next_shifttmpl else "—"

        # Determine status
        if session:
            status = session.status
            if session.status in ("present", "on_time", "early_arrival"):
                status = "present"
            elif session.status == "late":
                status = "late"
        else:
            # Check if on leave
            on_leave = (await db.execute(
                select(LeaveRequest).where(
                    and_(
                        LeaveRequest.employee_id == emp.id,
                        LeaveRequest.status == LeaveStatus.APPROVED,
                        LeaveRequest.start_date <= today,
                        LeaveRequest.end_date >= today,
                    )
                ).limit(1)
            )).scalar_one_or_none()
            if on_leave:
                status = "on_leave"
            elif shift_label == "OFF":
                status = "off_duty"
            else:
                # Check if checkin window has passed
                if shifttmpl and shifttmpl.checkin_window_end:
                    window_end = datetime.combine(today, shifttmpl.checkin_window_end)
                    if datetime.now() > window_end:
                        status = "absent"
                    else:
                        status = "off_duty"
                else:
                    status = "off_duty"

        # Apply filters
        if status_filter:
            if status_filter == "on_duty" and status not in ("present", "late"):
                continue
            if status_filter == "off_duty" and status != "off_duty":
                continue
            if status_filter == "present" and status != "present":
                continue
            if status_filter == "absent" and status != "absent":
                continue
            if status_filter == "late" and status != "late":
                continue
            if status_filter == "on_leave" and status != "on_leave":
                continue
        if shift_filter:
            if shift_filter == "day" and shift_label != "D":
                continue
            if shift_filter == "night" and shift_label != "N":
                continue

        employee_rows.append({
            "id": str(emp.id),
            "employee_code": emp.employee_code,
            "full_name": emp.full_name,
            "position": emp.position,
            "department_name": dept.name,
            "current_shift": shift_name,
            "shift_label": shift_label,
            "next_shift": next_shift_name,
            "next_shift_label": next_shift_label,
            "status": status,
            "check_in": session.check_in.isoformat() if session and session.check_in else None,
            "check_out": session.check_out.isoformat() if session and session.check_out else None,
            "late_minutes": float(session.late_minutes or 0) if session else 0,
            "shift_protocol_name": emp_protocol_name,
            "has_individual_override": has_individual_override,
        })

    return {
        "department": {
            "id": str(dept.id),
            "name": dept.name,
            "code": dept.code,
            "office_name": dept.office.name if dept.office else None,
            "shift_protocol_name": dept.shift_protocol.name if dept.shift_protocol else None,
            "protocol_type": dept.shift_protocol.protocol_type.value if dept.shift_protocol else "fixed",
            "head_name": dept.head_name,
        },
        "employees": employee_rows,
        "total": len(employee_rows),
        "date": str(today),
    }


# ── Monthly Roster Grid ──────────────────────────────────────

@router.get("/departments/{dept_id}/roster", dependencies=[Depends(PermissionChecker("employee:view"))])
async def department_roster(
    dept_id: UUID,
    year: int = Query(..., ge=2024, le=2030),
    month: int = Query(..., ge=1, le=12),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Monthly roster grid for a department."""
    days = _calendar_days(year, month)
    today = date.today()

    dept = (await db.execute(
        select(Department).where(Department.id == dept_id)
    )).scalar_one_or_none()
    if not dept:
        raise HTTPException(404, "Department not found")

    employees = (await db.execute(
        select(Employee).where(
            and_(Employee.department_id == dept_id, Employee.status == EmployeeStatus.ACTIVE)
        ).order_by(Employee.full_name)
    )).scalars().all()

    # Check approved leaves for this month
    leave_map: dict[UUID, list[tuple[date, date]]] = {}
    for emp in employees:
        leaves = (await db.execute(
            select(LeaveRequest).where(
                and_(
                    LeaveRequest.employee_id == emp.id,
                    LeaveRequest.status == LeaveStatus.APPROVED,
                    LeaveRequest.start_date <= days[-1],
                    LeaveRequest.end_date >= days[0],
                )
            )
        )).scalars().all()
        if leaves:
            leave_map[emp.id] = [(l.start_date, l.end_date) for l in leaves]

    roster_rows = []
    for emp in employees:
        daily = []
        for d in days:
            # Check leave first
            is_on_leave = False
            if emp.id in leave_map:
                for ls, le in leave_map[emp.id]:
                    if ls <= d <= le:
                        is_on_leave = True
                        break

            if is_on_leave:
                label = "LV"
            else:
                label = await _get_shift_label(db, emp.id, d)

            is_today = (d == today)
            daily.append({
                "date": str(d),
                "label": label,
                "is_today": is_today,
            })

        shifttmpl = await _resolve_shift_for_employee_on_date(db, emp.id, today)
        roster_rows.append({
            "employee_id": str(emp.id),
            "employee_code": emp.employee_code,
            "full_name": emp.full_name,
            "position": emp.position,
            "current_shift": shifttmpl.name if shifttmpl else "Unassigned",
            "daily": daily,
        })

    return {
        "department": {"id": str(dept.id), "name": dept.name},
        "year": year,
        "month": month,
        "days": [str(d) for d in days],
        "employees": roster_rows,
    }


# ── Employee Workforce Profile ────────────────────────────────

@router.get("/employees/{emp_id}/profile", dependencies=[Depends(PermissionChecker("employee:view"))])
async def employee_profile(
    emp_id: UUID,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Employee workforce profile with attendance summary and current assignment."""
    today = date.today()
    month_start = today.replace(day=1)

    emp = (await db.execute(
        select(Employee).options(
            joinedload(Employee.department),
            joinedload(Employee.shift_protocol),
        ).where(Employee.id == emp_id)
    )).scalar_one_or_none()
    if not emp:
        raise HTTPException(404, "Employee not found")

    # Department
    dept_name = emp.department.name if emp.department else "Unassigned"

    # Current shift
    current_shift = await _resolve_shift_for_employee_on_date(db, emp.id, today)
    current_shift_label = await _get_shift_label(db, emp.id, today)

    # Next shift
    tomorrow = today + timedelta(days=1)
    next_shift = await _resolve_shift_for_employee_on_date(db, emp.id, tomorrow)
    next_shift_label = await _get_shift_label(db, emp.id, tomorrow)

    # Next working day (scan forward up to 14 days)
    next_working_day = None
    next_working_shift = None
    for i in range(1, 15):
        check_date = today + timedelta(days=i)
        lbl = await _get_shift_label(db, emp.id, check_date)
        if lbl not in ("OFF", "—"):
            next_working_day = str(check_date)
            tmpl = await _resolve_shift_for_employee_on_date(db, emp.id, check_date)
            next_working_shift = tmpl.name if tmpl else None
            break

    # Monthly attendance stats
    sessions = (await db.execute(
        select(AttendanceSession).where(
            and_(
                AttendanceSession.employee_id == emp.id,
                AttendanceSession.date >= month_start,
                AttendanceSession.date <= today,
            )
        )
    )).scalars().all()

    present_count = sum(
        1 for s in sessions
        if s.status in ("present", "on_time", "early_arrival", "unscheduled_attendance")
    )
    late_count = sum(1 for s in sessions if s.status == "late")
    absent_count = sum(1 for s in sessions if s.status in ("absent", "missed_checkin"))
    overtime_hours = round(sum(float(s.overtime_minutes or 0) for s in sessions) / 60, 1)

    # Current assignment details
    assignment = (await db.execute(
        select(EmployeeShiftAssignment).where(
            EmployeeShiftAssignment.employee_id == emp.id
        ).order_by(EmployeeShiftAssignment.created_at.desc()).limit(1)
    )).scalar_one_or_none()

    roster_type = "unassigned"
    if assignment:
        roster_type = "rotating" if assignment.is_rotating else "fixed"

    return {
        "employee": {
            "id": str(emp.id),
            "employee_code": emp.employee_code,
            "full_name": emp.full_name,
            "email": emp.email,
            "phone": emp.phone,
            "position": emp.position,
            "status": emp.status.value if hasattr(emp.status, 'value') else str(emp.status),
            "department_id": str(emp.department_id) if emp.department_id else None,
            "department_name": dept_name,
        },
        "attendance_summary": {
            "present_this_month": present_count,
            "late_this_month": late_count,
            "absences_this_month": absent_count,
            "overtime_hours": overtime_hours,
            "total_sessions": len(sessions),
        },
        "current_assignment": {
            "department_name": dept_name,
            "roster_type": roster_type,
            "current_shift": current_shift.name if current_shift else "Unassigned",
            "current_shift_label": current_shift_label,
            "next_shift": next_shift.name if next_shift else "—",
            "next_shift_label": next_shift_label,
            "next_working_day": next_working_day,
            "next_working_shift": next_working_shift,
        },
    }


# ── Employee Monthly Calendar ─────────────────────────────────

@router.get("/employees/{emp_id}/calendar", dependencies=[Depends(PermissionChecker("employee:view"))])
async def employee_calendar(
    emp_id: UUID,
    year: int = Query(..., ge=2024, le=2030),
    month: int = Query(..., ge=1, le=12),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Monthly shift calendar for a single employee."""
    days = _calendar_days(year, month)
    today = date.today()

    emp = (await db.execute(
        select(Employee).where(Employee.id == emp_id)
    )).scalar_one_or_none()
    if not emp:
        raise HTTPException(404, "Employee not found")

    # Leaves
    leaves = (await db.execute(
        select(LeaveRequest).where(
            and_(
                LeaveRequest.employee_id == emp_id,
                LeaveRequest.status == LeaveStatus.APPROVED,
                LeaveRequest.start_date <= days[-1],
                LeaveRequest.end_date >= days[0],
            )
        )
    )).scalars().all()
    leave_ranges = [(l.start_date, l.end_date, l.leave_type.value) for l in leaves]

    # Sessions
    sessions = (await db.execute(
        select(AttendanceSession).where(
            and_(
                AttendanceSession.employee_id == emp_id,
                AttendanceSession.date >= days[0],
                AttendanceSession.date <= days[-1],
            )
        )
    )).scalars().all()
    session_map = {s.date: s for s in sessions}

    calendar_days = []
    for d in days:
        is_on_leave = False
        leave_type = None
        for ls, le, lt in leave_ranges:
            if ls <= d <= le:
                is_on_leave = True
                leave_type = lt
                break

        if is_on_leave:
            label = "LV"
        else:
            label = await _get_shift_label(db, emp_id, d)

        session = session_map.get(d)
        attendance_status = None
        if session:
            attendance_status = session.status

        calendar_days.append({
            "date": str(d),
            "day_of_week": d.strftime("%a"),
            "label": label,
            "is_today": d == today,
            "is_past": d < today,
            "attendance_status": attendance_status,
            "leave_type": leave_type,
            "check_in": session.check_in.isoformat() if session and session.check_in else None,
            "check_out": session.check_out.isoformat() if session and session.check_out else None,
        })

    return {
        "employee": {
            "id": str(emp.id),
            "employee_code": emp.employee_code,
            "full_name": emp.full_name,
        },
        "year": year,
        "month": month,
        "calendar": calendar_days,
    }


# ── Shift Change ──────────────────────────────────────────────

@router.post("/shift-change", dependencies=[Depends(PermissionChecker("shift:assign"))])
async def shift_change(
    req: ShiftChangeRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Create a time-bounded shift override for an employee."""
    # Validate employee
    emp = (await db.execute(
        select(Employee).where(Employee.id == req.employee_id)
    )).scalar_one_or_none()
    if not emp:
        raise HTTPException(404, "Employee not found")

    # Validate shift template
    tmpl = (await db.execute(
        select(ShiftTemplate).where(ShiftTemplate.id == req.shift_template_id)
    )).scalar_one_or_none()
    if not tmpl:
        raise HTTPException(404, "Shift template not found")

    if req.end_date < req.start_date:
        raise HTTPException(400, "end_date must be >= start_date")

    # Get old shift for audit
    old_shift = await _resolve_shift_for_employee_on_date(db, req.employee_id, req.start_date)
    old_shift_name = old_shift.name if old_shift else "Unassigned"

    # Create override
    override = EmployeeShiftOverride(
        employee_id=req.employee_id,
        shift_template_id=req.shift_template_id,
        start_date=req.start_date,
        end_date=req.end_date,
        reason=req.reason,
        created_by=current_user.id,
    )
    db.add(override)

    # Audit log
    audit = AuditLog(
        user_id=current_user.id,
        action="shift_change",
        entity_type="employee_shift_override",
        entity_id=str(req.employee_id),
        description=f"Shift changed from {old_shift_name} to {tmpl.name} ({req.start_date} to {req.end_date})",
        details={
            "employee_id": str(req.employee_id),
            "old_shift": old_shift_name,
            "new_shift": tmpl.name,
            "start_date": str(req.start_date),
            "end_date": str(req.end_date),
            "reason": req.reason,
        },
    )
    db.add(audit)

    await db.commit()

    return {
        "message": "Shift change applied",
        "override_id": str(override.id),
        "employee": emp.full_name,
        "old_shift": old_shift_name,
        "new_shift": tmpl.name,
        "period": f"{req.start_date} to {req.end_date}",
    }


# ── Shift Swap ────────────────────────────────────────────────

@router.post("/shift-swap", dependencies=[Depends(PermissionChecker("shift:assign"))])
async def shift_swap(
    req: ShiftSwapRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Swap shifts between two employees for a specific date."""
    emp_a = (await db.execute(
        select(Employee).where(Employee.id == req.employee_a_id)
    )).scalar_one_or_none()
    emp_b = (await db.execute(
        select(Employee).where(Employee.id == req.employee_b_id)
    )).scalar_one_or_none()

    if not emp_a or not emp_b:
        raise HTTPException(404, "One or both employees not found")
    if req.employee_a_id == req.employee_b_id:
        raise HTTPException(400, "Cannot swap with yourself")

    shift_a = await _resolve_shift_for_employee_on_date(db, req.employee_a_id, req.swap_date)
    shift_b = await _resolve_shift_for_employee_on_date(db, req.employee_b_id, req.swap_date)

    # Create overrides: assign B's shift to A, and A's shift to B
    if shift_b:
        override_a = EmployeeShiftOverride(
            employee_id=req.employee_a_id,
            shift_template_id=shift_b.id,
            start_date=req.swap_date,
            end_date=req.swap_date,
            reason=f"Shift swap with {emp_b.full_name}: {req.reason or 'No reason'}",
            created_by=current_user.id,
        )
        db.add(override_a)

    if shift_a:
        override_b = EmployeeShiftOverride(
            employee_id=req.employee_b_id,
            shift_template_id=shift_a.id,
            start_date=req.swap_date,
            end_date=req.swap_date,
            reason=f"Shift swap with {emp_a.full_name}: {req.reason or 'No reason'}",
            created_by=current_user.id,
        )
        db.add(override_b)

    # Audit
    audit = AuditLog(
        user_id=current_user.id,
        action="shift_swap",
        entity_type="employee_shift_override",
        entity_id=f"{req.employee_a_id},{req.employee_b_id}",
        description=f"Shift swap: {emp_a.full_name} ({shift_a.name if shift_a else '—'}) ↔ {emp_b.full_name} ({shift_b.name if shift_b else '—'}) on {req.swap_date}",
        details={
            "employee_a": str(req.employee_a_id),
            "employee_b": str(req.employee_b_id),
            "swap_date": str(req.swap_date),
            "shift_a": shift_a.name if shift_a else None,
            "shift_b": shift_b.name if shift_b else None,
            "reason": req.reason,
        },
    )
    db.add(audit)

    await db.commit()

    return {
        "message": "Shift swap applied",
        "employee_a": emp_a.full_name,
        "employee_b": emp_b.full_name,
        "swap_date": str(req.swap_date),
    }


# ── Shift Coverage ────────────────────────────────────────────

@router.get("/coverage", dependencies=[Depends(PermissionChecker("employee:view"))])
async def shift_coverage(
    department_id: Optional[UUID] = Query(None),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Current shift coverage stats."""
    today = date.today()
    now = datetime.now()

    emp_query = select(Employee).where(Employee.status == EmployeeStatus.ACTIVE)
    if department_id:
        emp_query = emp_query.where(Employee.department_id == department_id)

    employees = (await db.execute(emp_query)).scalars().all()

    day_count = 0
    night_count = 0
    off_count = 0
    total = len(employees)

    for emp in employees:
        lbl = await _get_shift_label(db, emp.id, today)
        if lbl == "D":
            day_count += 1
        elif lbl == "N":
            night_count += 1
        else:
            off_count += 1

    # Present counts
    present_today = (await db.execute(
        select(func.count()).select_from(AttendanceSession)
        .join(Employee, Employee.id == AttendanceSession.employee_id)
        .where(
            and_(
                AttendanceSession.date == today,
                AttendanceSession.status.in_(["present", "on_time", "early_arrival"]),
                *([Employee.department_id == department_id] if department_id else []),
            )
        )
    )).scalar_one()

    return {
        "date": str(today),
        "time": now.strftime("%H:%M"),
        "total_employees": total,
        "day_shift": day_count,
        "night_shift": night_count,
        "off_duty": off_count,
        "present_now": present_today,
        "day_coverage": round(day_count / max(total, 1) * 100, 1),
        "night_coverage": round(night_count / max(total, 1) * 100, 1),
    }


# ── Upcoming Changes ─────────────────────────────────────────

@router.get("/upcoming-changes", dependencies=[Depends(PermissionChecker("employee:view"))])
async def upcoming_changes(
    days_ahead: int = Query(7, ge=1, le=30),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Upcoming shift changes and leave returns."""
    today = date.today()
    end_date = today + timedelta(days=days_ahead)

    # Shift overrides in the window
    overrides = (await db.execute(
        select(EmployeeShiftOverride)
        .options(joinedload(EmployeeShiftOverride.shift_template))
        .where(
            and_(
                EmployeeShiftOverride.start_date >= today,
                EmployeeShiftOverride.start_date <= end_date,
            )
        ).order_by(EmployeeShiftOverride.start_date)
    )).scalars().all()

    changes = []
    for o in overrides:
        emp = (await db.execute(
            select(Employee).where(Employee.id == o.employee_id)
        )).scalar_one_or_none()
        if emp:
            dept = None
            if emp.department_id:
                dept = (await db.execute(
                    select(Department).where(Department.id == emp.department_id)
                )).scalar_one_or_none()
            changes.append({
                "type": "shift_change",
                "employee_id": str(emp.id),
                "employee_name": emp.full_name,
                "department_name": dept.name if dept else "Unassigned",
                "new_shift": o.shift_template.name if o.shift_template else "Unknown",
                "effective_date": str(o.start_date),
                "reason": o.reason,
            })

    # Leave returns in the window
    returning_leaves = (await db.execute(
        select(LeaveRequest).where(
            and_(
                LeaveRequest.status == LeaveStatus.APPROVED,
                LeaveRequest.end_date >= today,
                LeaveRequest.end_date <= end_date,
            )
        ).order_by(LeaveRequest.end_date)
    )).scalars().all()

    for lr in returning_leaves:
        emp = (await db.execute(
            select(Employee).where(Employee.id == lr.employee_id)
        )).scalar_one_or_none()
        if emp:
            dept = None
            if emp.department_id:
                dept = (await db.execute(
                    select(Department).where(Department.id == emp.department_id)
                )).scalar_one_or_none()
            changes.append({
                "type": "returning_from_leave",
                "employee_id": str(emp.id),
                "employee_name": emp.full_name,
                "department_name": dept.name if dept else "Unassigned",
                "leave_type": lr.leave_type.value,
                "return_date": str(lr.end_date + timedelta(days=1)),
            })

    # Sort by effective date
    changes.sort(key=lambda x: x.get("effective_date") or x.get("return_date", ""))

    return {"changes": changes, "total": len(changes), "period": f"{today} to {end_date}"}


# ── Roster Export ─────────────────────────────────────────────

@router.post("/roster/export", dependencies=[Depends(PermissionChecker("report:export"))])
async def export_roster(
    req: RosterExportRequest,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Export department roster as Excel or CSV."""
    days = _calendar_days(req.year, req.month)

    dept = (await db.execute(
        select(Department).where(Department.id == req.department_id)
    )).scalar_one_or_none()
    if not dept:
        raise HTTPException(404, "Department not found")

    employees = (await db.execute(
        select(Employee).where(
            and_(Employee.department_id == req.department_id, Employee.status == EmployeeStatus.ACTIVE)
        ).order_by(Employee.full_name)
    )).scalars().all()

    # Build roster data
    roster_data = []
    for emp in employees:
        row = {
            "Employee ID": emp.employee_code,
            "Name": emp.full_name,
            "Designation": emp.position or "",
        }
        for d in days:
            label = await _get_shift_label(db, emp.id, d)
            row[d.strftime("%d")] = label
        roster_data.append(row)

    if req.format == "xlsx":
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{dept.name} - {req.month}/{req.year}"

        # Header
        headers = ["Employee ID", "Name", "Designation"] + [d.strftime("%d %a") for d in days]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        # Data
        for row_idx, row_data in enumerate(roster_data, 2):
            for col_idx, key in enumerate(row_data.keys(), 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data[key])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"roster_{dept.code}_{req.year}_{req.month:02d}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    else:
        # CSV
        import csv
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=[
            "Employee ID", "Name", "Designation"
        ] + [d.strftime("%d") for d in days])
        writer.writeheader()
        for row_data in roster_data:
            writer.writerow(row_data)

        content = buf.getvalue().encode("utf-8")
        filename = f"roster_{dept.code}_{req.year}_{req.month:02d}.csv"
        return StreamingResponse(
            io.BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
