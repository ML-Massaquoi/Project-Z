"""
Project Z - Daily Reports API
Generate, retrieve, and export daily attendance reports.

Report generation logic:
  - For each active employee on a given date
  - First scan = check-in, last scan = check-out
  - Calculates late minutes, overtime, duration
  - Detects absent, on-leave, off-duty employees
  - Saves to daily_reports + daily_report_lines (idempotent)
"""

import io
import csv
import logging
from datetime import date, datetime, time, timedelta, timezone
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import and_, func, select, delete
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.core.dependencies import get_current_user, PermissionChecker
from app.database.session import get_db
from app.models.attendance import AttendanceSession
from app.models.daily_report import DailyReport, DailyReportLine
from app.models.department import Department
from app.models.employee import Employee, EmployeeStatus
from app.models.leave_request import LeaveRequest, LeaveStatus
from app.models.scan_event import ScanEvent, ScanResult
from app.models.shift_override import EmployeeShiftOverride
from app.models.shift_assignment import EmployeeShiftAssignment
from app.models.shift_template import ShiftTemplate
from app.models.dept_shift_rule import DepartmentShiftRule

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/daily-reports", tags=["Daily Reports"])


# ── Schemas ────────────────────────────────────────────────────

class GenerateReportRequest(BaseModel):
    report_date: str  # YYYY-MM-DD
    department_id: Optional[str] = None


# ── Helpers ────────────────────────────────────────────────────

async def _resolve_shift_for_emp_on_date(
    db: AsyncSession, employee_id: UUID, target_date: date
) -> Optional[ShiftTemplate]:
    """4-level shift resolution: override → assignment → dept rule → None."""
    # Level 1: Override
    override = (await db.execute(
        select(EmployeeShiftOverride).where(
            and_(
                EmployeeShiftOverride.employee_id == employee_id,
                EmployeeShiftOverride.start_date <= target_date,
                EmployeeShiftOverride.end_date >= target_date,
            )
        ).limit(1)
    )).scalar_one_or_none()
    if override:
        return (await db.execute(
            select(ShiftTemplate).where(ShiftTemplate.id == override.shift_template_id)
        )).scalar_one_or_none()

    # Level 2: Assignment
    assignment = (await db.execute(
        select(EmployeeShiftAssignment).where(
            EmployeeShiftAssignment.employee_id == employee_id
        ).limit(1)
    )).scalar_one_or_none()
    if assignment:
        tmpl_id = assignment.resolve_template_id_for_date(target_date)
        if tmpl_id:
            return (await db.execute(
                select(ShiftTemplate).where(ShiftTemplate.id == tmpl_id)
            )).scalar_one_or_none()

    # Level 3: Department rule
    emp = (await db.execute(
        select(Employee).where(Employee.id == employee_id)
    )).scalar_one_or_none()
    if not emp or not emp.department_id:
        return None

    dept_rule = (await db.execute(
        select(DepartmentShiftRule).where(
            and_(
                DepartmentShiftRule.department_id == emp.department_id,
                DepartmentShiftRule.effective_from <= target_date,
                (DepartmentShiftRule.effective_to.is_(None) | (DepartmentShiftRule.effective_to >= target_date)),
            )
        ).order_by(DepartmentShiftRule.effective_from.desc()).limit(1)
    )).scalar_one_or_none()
    if dept_rule:
        return (await db.execute(
            select(ShiftTemplate).where(ShiftTemplate.id == dept_rule.shift_template_id)
        )).scalar_one_or_none()

    return None


def _calculate_late_minutes(shift_start: time, check_in: datetime, grace_minutes: int = 0) -> float:
    """Calculate how many minutes late an employee is."""
    if not shift_start or not check_in:
        return 0.0
    shift_start_dt = datetime.combine(check_in.date(), shift_start, tzinfo=check_in.tzinfo or timezone.utc)
    grace_delta = timedelta(minutes=grace_minutes)
    late = (check_in - shift_start_dt - grace_delta).total_seconds() / 60
    return max(0.0, late)


def _calculate_overtime_minutes(shift_end: time, check_out: datetime) -> float:
    """Calculate overtime beyond shift end."""
    if not shift_end or not check_out:
        return 0.0
    shift_end_dt = datetime.combine(check_out.date(), shift_end, tzinfo=check_out.tzinfo or timezone.utc)
    overtime = (check_out - shift_end_dt).total_seconds() / 60
    return max(0.0, overtime)


def _calculate_early_departure(shift_end: time, check_out: datetime) -> float:
    """Calculate how many minutes early an employee left."""
    if not shift_end or not check_out:
        return 0.0
    shift_end_dt = datetime.combine(check_out.date(), shift_end, tzinfo=check_out.tzinfo or timezone.utc)
    early = (shift_end_dt - check_out).total_seconds() / 60
    return max(0.0, early)


def _format_datetime(dt: Optional[datetime]) -> Optional[str]:
    return dt.isoformat() if dt else None


def _format_time(t: Optional[time]) -> Optional[str]:
    return t.isoformat() if t else None


# ── Generate Report ────────────────────────────────────────────

@router.post("/generate", dependencies=[Depends(PermissionChecker("report:view"))])
async def generate_daily_report(
    req: GenerateReportRequest,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    """
    Generate a daily attendance report for a specific date.
    For each active employee: first scan = check-in, last scan = check-out.
    Saves to daily_reports + daily_report_lines (replaces existing for same date+dept).
    """
    target_date = date.fromisoformat(req.report_date)

    # Get departments to generate for
    dept_query = select(Department).where(Department.is_active == True)
    if req.department_id:
        dept_query = dept_query.where(Department.id == UUID(req.department_id))
    departments = (await db.execute(dept_query)).scalars().all()

    if not departments:
        raise HTTPException(404, "No active departments found")

    now = datetime.now(timezone.utc)
    generated_reports = []

    for dept in departments:
        # Delete existing report for this date+dept (idempotent re-generation)
        existing = (await db.execute(
            select(DailyReport).where(
                and_(DailyReport.report_date == target_date, DailyReport.department_id == dept.id)
            )
        )).scalar_one_or_none()
        if existing:
            await db.delete(existing)
            await db.flush()

        # Get all active employees in this department
        employees = (await db.execute(
            select(Employee).where(
                and_(Employee.department_id == dept.id, Employee.status == EmployeeStatus.ACTIVE)
            ).order_by(Employee.full_name)
        )).scalars().all()

        if not employees:
            continue

        # Create report header
        report = DailyReport(
            report_date=target_date,
            department_id=dept.id,
            department_name=dept.name,
            total_expected=len(employees),
            generated_at=now,
            generated_by=user.id if hasattr(user, 'id') else None,
            is_final=False,
        )
        db.add(report)
        await db.flush()

        total_present = 0
        total_late = 0
        total_absent = 0
        total_on_leave = 0
        total_overtime = 0
        total_early_departure = 0

        for emp in employees:
            # Check if on leave
            on_leave = (await db.execute(
                select(LeaveRequest).where(
                    and_(
                        LeaveRequest.employee_id == emp.id,
                        LeaveRequest.status == LeaveStatus.APPROVED,
                        LeaveRequest.start_date <= target_date,
                        LeaveRequest.end_date >= target_date,
                    )
                ).limit(1)
            )).scalar_one_or_none()

            if on_leave:
                total_on_leave += 1
                line = DailyReportLine(
                    report_id=report.id,
                    employee_id=emp.id,
                    employee_code=emp.employee_code,
                    employee_name=emp.full_name,
                    department_name=dept.name,
                    position=emp.position,
                    status="on_leave",
                    late_minutes=0,
                    overtime_minutes=0,
                    early_departure_minutes=0,
                    duration_minutes=0,
                )
                db.add(line)
                continue

            # Resolve shift template
            shifttmpl = await _resolve_shift_for_emp_on_date(db, emp.id, target_date)

            shift_name = shifttmpl.name if shifttmpl else None
            shift_start = shifttmpl.start_time if shifttmpl else None
            shift_end = shifttmpl.end_time if shifttmpl else None
            grace_minutes = shifttmpl.grace_period_minutes if shifttmpl else 0

            # Get all scans for this employee in a 48h window around target_date
            # This captures overnight shifts: check-in on target_date, check-out next day
            # Also captures early-morning check-ins from previous night's shift
            scan_window_start = datetime.combine(target_date - timedelta(days=1), time.min, tzinfo=timezone.utc)
            scan_window_end = datetime.combine(target_date + timedelta(days=2), time.min, tzinfo=timezone.utc)

            scans = (await db.execute(
                select(ScanEvent).where(
                    and_(
                        ScanEvent.employee_id == emp.id,
                        ScanEvent.scan_timestamp >= scan_window_start,
                        ScanEvent.scan_timestamp < scan_window_end,
                        ScanEvent.scan_result == ScanResult.SUCCESSFUL,
                    )
                ).order_by(ScanEvent.scan_timestamp)
            )).scalars().all()

            # ── Classify scans for this shift ─────────────────────────
            # A scan belongs to this day's shift if:
            #   - Normal shift (start < end): scan is on target_date
            #   - Overnight shift (start > end): scan is between shift_start on target_date and shift_end on target_date+1
            #   - Check-in from previous night: scan is on target_date-1 evening after shift_start, for today's early shift
            if shift_start and shift_end and shift_start > shift_end:
                # Overnight shift: e.g., 20:00 → 08:00
                # Check-in window: target_date at shift_start through target_date+1 at shift_end
                checkin_lower = datetime.combine(target_date, shift_start, tzinfo=timezone.utc)
                checkout_upper = datetime.combine(target_date + timedelta(days=1), shift_end, tzinfo=timezone.utc)
                relevant_scans = [s for s in scans if checkin_lower <= s.scan_timestamp <= checkout_upper]
            elif shift_start and shift_end:
                # Normal shift: e.g., 08:00 → 17:00
                # Check-in on target_date, check-out can be same day or slightly past midnight
                checkin_lower = datetime.combine(target_date, time.min, tzinfo=timezone.utc)
                checkout_upper = datetime.combine(target_date + timedelta(days=1), time(3, 0), tzinfo=timezone.utc)  # 3 AM grace
                relevant_scans = [s for s in scans if checkin_lower <= s.scan_timestamp <= checkout_upper]
            else:
                # No shift defined — just use target_date scans
                checkin_lower = datetime.combine(target_date, time.min, tzinfo=timezone.utc)
                checkout_upper = datetime.combine(target_date + timedelta(days=1), time.min, tzinfo=timezone.utc)
                relevant_scans = [s for s in scans if checkin_lower <= s.scan_timestamp <= checkout_upper]

            if not relevant_scans:
                # No scans → absent
                total_absent += 1
                line = DailyReportLine(
                    report_id=report.id,
                    employee_id=emp.id,
                    employee_code=emp.employee_code,
                    employee_name=emp.full_name,
                    department_name=dept.name,
                    position=emp.position,
                    shift_name=shift_name,
                    shift_start=shift_start,
                    shift_end=shift_end,
                    total_scans=0,
                    status="absent",
                    late_minutes=0,
                    overtime_minutes=0,
                    early_departure_minutes=0,
                    duration_minutes=0,
                )
                db.add(line)
                continue

            # First scan = check-in, last scan = check-out
            first_scan = relevant_scans[0].scan_timestamp
            last_scan = relevant_scans[-1].scan_timestamp
            total_scans = len(relevant_scans)

            # Check-in device
            check_in_device = relevant_scans[0].device_name
            check_out_device = relevant_scans[-1].device_name if total_scans > 1 else None

            # Calculate metrics
            late_minutes = _calculate_late_minutes(shift_start, first_scan, grace_minutes) if shift_start else 0
            duration = (last_scan - first_scan).total_seconds() / 60

            # Overtime: handle overnight shifts (shift_end < shift_start means crosses midnight)
            overtime_minutes = 0.0
            early_departure = 0.0
            if shift_start and shift_end:
                if shift_end > shift_start:
                    # Normal shift — overtime after shift_end on same day
                    overtime_minutes = _calculate_overtime_minutes(shift_end, last_scan)
                    early_departure = _calculate_early_departure(shift_end, last_scan) if total_scans > 1 else 0
                else:
                    # Overnight shift — overtime after shift_end on NEXT day
                    checkout_dt_next_day = datetime.combine(target_date + timedelta(days=1), shift_end, tzinfo=timezone.utc)
                    overtime_minutes = max(0.0, (last_scan - checkout_dt_next_day).total_seconds() / 60)

            # Determine status
            if shift_start:
                if late_minutes > 0:
                    status = "late"
                    total_late += 1
                else:
                    status = "on_time"
            else:
                status = "present"

            total_present += 1
            if overtime_minutes > 0:
                total_overtime += 1
            if early_departure > 0:
                total_early_departure += 1

            line = DailyReportLine(
                report_id=report.id,
                employee_id=emp.id,
                employee_code=emp.employee_code,
                employee_name=emp.full_name,
                department_name=dept.name,
                position=emp.position,
                shift_name=shift_name,
                shift_start=shift_start,
                shift_end=shift_end,
                first_scan=first_scan,
                last_scan=last_scan,
                total_scans=total_scans,
                check_in=first_scan,
                check_out=last_scan,
                late_minutes=late_minutes,
                overtime_minutes=overtime_minutes,
                early_departure_minutes=early_departure,
                duration_minutes=round(duration, 1),
                status=status,
                check_in_device=check_in_device,
                check_out_device=check_out_device,
            )
            db.add(line)

        # Update report totals
        report.total_present = total_present
        report.total_late = total_late
        report.total_absent = total_absent
        report.total_on_leave = total_on_leave
        report.total_overtime = total_overtime
        report.total_early_departure = total_early_departure

        generated_reports.append({
            "report_id": str(report.id),
            "department": dept.name,
            "date": str(target_date),
            "total_expected": report.total_expected,
            "total_present": total_present,
            "total_late": total_late,
            "total_absent": total_absent,
            "total_on_leave": total_on_leave,
        })

    await db.commit()

    return {
        "message": f"Generated {len(generated_reports)} daily report(s) for {target_date}",
        "reports": generated_reports,
    }


# ── List Reports ───────────────────────────────────────────────

@router.get("/list", dependencies=[Depends(PermissionChecker("report:view"))])
async def list_daily_reports(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    department_id: Optional[UUID] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """List all generated daily reports with optional filters."""
    query = select(DailyReport).order_by(DailyReport.report_date.desc())

    if start_date:
        query = query.where(DailyReport.report_date >= date.fromisoformat(start_date))
    if end_date:
        query = query.where(DailyReport.report_date <= date.fromisoformat(end_date))
    if department_id:
        query = query.where(DailyReport.department_id == department_id)

    # Count
    count_query = select(func.count()).select_from(DailyReport)
    if start_date:
        count_query = count_query.where(DailyReport.report_date >= date.fromisoformat(start_date))
    if end_date:
        count_query = count_query.where(DailyReport.report_date <= date.fromisoformat(end_date))
    if department_id:
        count_query = count_query.where(DailyReport.department_id == department_id)
    total = (await db.execute(count_query)).scalar_one()

    # Paginate
    query = query.offset((page - 1) * per_page).limit(per_page)
    result = await db.execute(query)
    reports = result.scalars().all()

    return {
        "items": [
            {
                "id": str(r.id),
                "report_date": str(r.report_date),
                "department_id": str(r.department_id),
                "department_name": r.department_name,
                "total_expected": r.total_expected,
                "total_present": r.total_present,
                "total_late": r.total_late,
                "total_absent": r.total_absent,
                "total_on_leave": r.total_on_leave,
                "total_overtime": r.total_overtime,
                "total_early_departure": r.total_early_departure,
                "generated_at": r.generated_at.isoformat() if r.generated_at else None,
                "is_final": r.is_final,
            }
            for r in reports
        ],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
    }


# ── Get Report Detail ──────────────────────────────────────────

@router.get("/{report_id}", dependencies=[Depends(PermissionChecker("report:view"))])
async def get_daily_report_detail(
    report_id: UUID,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Get full daily report with all employee lines."""
    report = (await db.execute(
        select(DailyReport).where(DailyReport.id == report_id)
    )).scalar_one_or_none()
    if not report:
        raise HTTPException(404, "Report not found")

    lines = (await db.execute(
        select(DailyReportLine).where(DailyReportLine.report_id == report_id)
        .order_by(DailyReportLine.status, DailyReportLine.employee_name)
    )).scalars().all()

    return {
        "report": {
            "id": str(report.id),
            "report_date": str(report.report_date),
            "department_id": str(report.department_id),
            "department_name": report.department_name,
            "total_expected": report.total_expected,
            "total_present": report.total_present,
            "total_late": report.total_late,
            "total_absent": report.total_absent,
            "total_on_leave": report.total_on_leave,
            "total_overtime": report.total_overtime,
            "total_early_departure": report.total_early_departure,
            "generated_at": report.generated_at.isoformat() if report.generated_at else None,
            "is_final": report.is_final,
        },
        "lines": [
            {
                "id": str(l.id),
                "employee_id": str(l.employee_id),
                "employee_code": l.employee_code,
                "employee_name": l.employee_name,
                "department_name": l.department_name,
                "position": l.position,
                "shift_name": l.shift_name,
                "shift_start": _format_time(l.shift_start),
                "shift_end": _format_time(l.shift_end),
                "first_scan": _format_datetime(l.first_scan),
                "last_scan": _format_datetime(l.last_scan),
                "total_scans": l.total_scans,
                "check_in": _format_datetime(l.check_in),
                "check_out": _format_datetime(l.check_out),
                "late_minutes": l.late_minutes,
                "overtime_minutes": l.overtime_minutes,
                "early_departure_minutes": l.early_departure_minutes,
                "duration_minutes": l.duration_minutes,
                "status": l.status,
                "check_in_device": l.check_in_device,
                "check_out_device": l.check_out_device,
            }
            for l in lines
        ],
    }


# ── Get Report by Date + Department ────────────────────────────

@router.get("/by-date/{report_date}", dependencies=[Depends(PermissionChecker("report:view"))])
async def get_report_by_date(
    report_date: str,
    department_id: Optional[UUID] = Query(None),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Get report(s) for a specific date. If dept_id provided, return that specific report. Otherwise all depts."""
    target = date.fromisoformat(report_date)
    query = select(DailyReport).where(DailyReport.report_date == target)
    if department_id:
        query = query.where(DailyReport.department_id == department_id)
    query = query.order_by(DailyReport.department_name)

    reports = (await db.execute(query)).scalars().all()
    if not reports:
        raise HTTPException(404, f"No reports found for {report_date}")

    result = []
    for r in reports:
        lines = (await db.execute(
            select(DailyReportLine).where(DailyReportLine.report_id == r.id)
            .order_by(DailyReportLine.status, DailyReportLine.employee_name)
        )).scalars().all()
        result.append({
            "report": {
                "id": str(r.id),
                "report_date": str(r.report_date),
                "department_id": str(r.department_id),
                "department_name": r.department_name,
                "total_expected": r.total_expected,
                "total_present": r.total_present,
                "total_late": r.total_late,
                "total_absent": r.total_absent,
                "total_on_leave": r.total_on_leave,
                "total_overtime": r.total_overtime,
                "total_early_departure": r.total_early_departure,
                "generated_at": r.generated_at.isoformat() if r.generated_at else None,
            },
            "lines": [
                {
                    "id": str(l.id),
                    "employee_id": str(l.employee_id),
                    "employee_code": l.employee_code,
                    "employee_name": l.employee_name,
                    "department_name": l.department_name,
                    "position": l.position,
                    "shift_name": l.shift_name,
                    "shift_start": _format_time(l.shift_start),
                    "shift_end": _format_time(l.shift_end),
                    "first_scan": _format_datetime(l.first_scan),
                    "last_scan": _format_datetime(l.last_scan),
                    "total_scans": l.total_scans,
                    "check_in": _format_datetime(l.check_in),
                    "check_out": _format_datetime(l.check_out),
                    "late_minutes": l.late_minutes,
                    "overtime_minutes": l.overtime_minutes,
                    "early_departure_minutes": l.early_departure_minutes,
                    "duration_minutes": l.duration_minutes,
                    "status": l.status,
                    "check_in_device": l.check_in_device,
                    "check_out_device": l.check_out_device,
                }
                for l in lines
            ],
        })

    return {"date": report_date, "reports": result}


# ── Export Report ──────────────────────────────────────────────

@router.get("/{report_id}/export", dependencies=[Depends(PermissionChecker("report:export"))])
async def export_daily_report(
    report_id: UUID,
    format: str = Query("csv", description="csv | excel"),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Export a daily report as CSV or Excel."""
    report = (await db.execute(
        select(DailyReport).where(DailyReport.id == report_id)
    )).scalar_one_or_none()
    if not report:
        raise HTTPException(404, "Report not found")

    lines = (await db.execute(
        select(DailyReportLine).where(DailyReportLine.report_id == report_id)
        .order_by(DailyReportLine.employee_name)
    )).scalars().all()

    rows = []
    for l in lines:
        rows.append({
            "Employee Code": l.employee_code,
            "Employee Name": l.employee_name,
            "Department": l.department_name,
            "Position": l.position or "",
            "Shift": l.shift_name or "",
            "Shift Start": _format_time(l.shift_start) or "",
            "Shift End": _format_time(l.shift_end) or "",
            "Check In": l.first_scan.strftime("%H:%M:%S") if l.first_scan else "",
            "Check Out": l.last_scan.strftime("%H:%M:%S") if l.last_scan else "",
            "Duration (min)": round(l.duration_minutes, 1),
            "Late (min)": round(l.late_minutes, 1),
            "Overtime (min)": round(l.overtime_minutes, 1),
            "Early Departure (min)": round(l.early_departure_minutes, 1),
            "Scans": l.total_scans,
            "Status": l.status.replace("_", " ").title(),
            "Check-in Device": l.check_in_device or "",
            "Check-out Device": l.check_out_device or "",
        })

    filename = f"daily_report_{report.report_date}_{report.department_name}"

    if format == "excel":
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Daily Report {report.report_date}"

            # Header row
            headers = list(rows[0].keys()) if rows else []
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = openpyxl.styles.Font(bold=True)

            # Data rows
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, key in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data[key])

            # Auto-width
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
            )
        except ImportError:
            raise HTTPException(500, "Excel export requires openpyxl")

    # CSV fallback
    if not rows:
        raise HTTPException(404, "No data to export")

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=rows[0].keys())
    writer.writeheader()
    writer.writerows(rows)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
    )


# ── Delete Report ──────────────────────────────────────────────

@router.delete("/{report_id}", dependencies=[Depends(PermissionChecker("report:view"))])
async def delete_daily_report(
    report_id: UUID,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Delete a daily report and all its lines."""
    report = (await db.execute(
        select(DailyReport).where(DailyReport.id == report_id)
    )).scalar_one_or_none()
    if not report:
        raise HTTPException(404, "Report not found")

    await db.delete(report)
    await db.commit()

    return {"message": f"Deleted report for {report.report_date} - {report.department_name}"}
