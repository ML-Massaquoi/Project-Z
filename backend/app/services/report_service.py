"""
Project Z - Report Service
Generate attendance reports in PDF, Excel, and CSV formats.
"""

import io
import csv
from datetime import date, datetime
from typing import Optional
from uuid import UUID

from sqlalchemy import and_, select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.models.attendance import AttendanceSession
from app.models.employee import Employee
from app.models.department import Department


class ReportService:
    def __init__(self, session: AsyncSession):
        self.session = session

    async def generate_attendance_report(
        self,
        start_date: date,
        end_date: date,
        department_id: Optional[UUID] = None,
        format: str = "csv",
    ) -> tuple[bytes, str, str]:
        """
        Generate attendance report.
        Returns (content_bytes, filename, content_type).
        """
        # Fetch data
        query = (
            select(AttendanceSession)
            .options(joinedload(AttendanceSession.employee))
            .where(
                and_(
                    AttendanceSession.date >= start_date,
                    AttendanceSession.date <= end_date,
                )
            )
            .order_by(AttendanceSession.date, AttendanceSession.check_in)
        )
        if department_id:
            query = query.join(Employee).where(
                Employee.department_id == department_id
            )

        result = await self.session.execute(query)
        sessions = result.unique().scalars().all()

        if format == "csv":
            return self._generate_csv(sessions, start_date, end_date)
        elif format == "excel":
            return self._generate_excel(sessions, start_date, end_date)
        elif format == "pdf":
            return self._generate_pdf(sessions, start_date, end_date)
        else:
            return self._generate_csv(sessions, start_date, end_date)

    def _generate_csv(self, sessions, start_date, end_date):
        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow([
            "Date", "Employee Code", "Employee Name", "Check In",
            "Check Out", "Duration (min)", "Status", "Late (min)",
            "Overtime (min)"
        ])

        for s in sessions:
            emp_name = s.employee.full_name if s.employee else "Unknown"
            emp_code = s.employee.employee_code if s.employee else "N/A"
            writer.writerow([
                str(s.date),
                emp_code,
                emp_name,
                s.check_in.strftime("%H:%M:%S") if s.check_in else "",
                s.check_out.strftime("%H:%M:%S") if s.check_out else "",
                round(s.duration_minutes, 1) if s.duration_minutes else "",
                s.status.value if hasattr(s.status, 'value') else str(s.status),
                s.late_minutes or 0,
                s.overtime_minutes or 0,
            ])

        content = output.getvalue().encode("utf-8")
        filename = f"attendance_report_{start_date}_{end_date}.csv"
        return content, filename, "text/csv"

    def _generate_excel(self, sessions, start_date, end_date):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # Title
        ws.merge_cells('A1:I1')
        ws['A1'] = f"Attendance Report — {start_date} to {end_date}"
        ws['A1'].font = Font(bold=True, size=14, color="2563EB")

        ws.merge_cells('A2:I2')
        ws['A2'] = "Freetown International Airport"
        ws['A2'].font = Font(size=11, color="666666")

        # Headers
        headers = [
            "Date", "Employee Code", "Employee Name", "Check In",
            "Check Out", "Duration (min)", "Status", "Late (min)",
            "Overtime (min)"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data
        for row_idx, s in enumerate(sessions, 5):
            emp_name = s.employee.full_name if s.employee else "Unknown"
            emp_code = s.employee.employee_code if s.employee else "N/A"
            row_data = [
                str(s.date),
                emp_code,
                emp_name,
                s.check_in.strftime("%H:%M:%S") if s.check_in else "",
                s.check_out.strftime("%H:%M:%S") if s.check_out else "",
                round(s.duration_minutes, 1) if s.duration_minutes else "",
                s.status.value if hasattr(s.status, 'value') else str(s.status),
                s.late_minutes or 0,
                s.overtime_minutes or 0,
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()
        filename = f"attendance_report_{start_date}_{end_date}.xlsx"
        return content, filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def _generate_pdf(self, sessions, start_date, end_date):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()

        # Title
        elements.append(Paragraph(
            f"Attendance Report — {start_date} to {end_date}",
            styles['Title']
        ))
        elements.append(Paragraph("Freetown International Airport", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Table data
        data = [[
            "Date", "Code", "Name", "Check In", "Check Out",
            "Duration", "Status", "Late", "OT"
        ]]
        for s in sessions:
            emp_name = s.employee.full_name if s.employee else "Unknown"
            emp_code = s.employee.employee_code if s.employee else "N/A"
            data.append([
                str(s.date),
                emp_code,
                emp_name,
                s.check_in.strftime("%H:%M") if s.check_in else "",
                s.check_out.strftime("%H:%M") if s.check_out else "",
                f"{round(s.duration_minutes, 0)}" if s.duration_minutes else "",
                s.status.value if hasattr(s.status, 'value') else str(s.status),
                str(s.late_minutes or 0),
                str(s.overtime_minutes or 0),
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ]))
        elements.append(table)

        doc.build(elements)
        content = buffer.getvalue()
        filename = f"attendance_report_{start_date}_{end_date}.pdf"
        return content, filename, "application/pdf"
